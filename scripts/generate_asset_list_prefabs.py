#!/usr/bin/env python3
"""
Scan Unity scenes/prefabs for PrefabInstance sources and fill Asset List.xlsx.

Only prefabs reachable from project game scenes (Assets/Scenes + Editor Build Settings)
are listed — not asset-pack demos or example scenes.

Requires: pip install openpyxl (or use project .tools_openpyxl: PYTHONPATH=.tools_openpyxl)
"""
from __future__ import annotations

import argparse
import os
import re
from collections import defaultdict, deque

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
GUID_RE = re.compile(r"^guid:\s*([a-f0-9]{32})", re.M)
SRC_PREFAB_RE = re.compile(r"m_SourcePrefab:.*guid:\s*([a-f0-9]{32})")
SKIP_DIR_PARTS = ("/Library/", "/.git/", "/.tools_openpyxl", "/node_modules/")

# Appended after auto-detected prefabs: planned UI / menu work not yet in the repo.
# Columns: Asset Type, File Name, Link, Path, Context, Description, References, Status,
# Responsible, Est, To Date, Comments  (ID assigned at export)
FUTURE_PLANNED_PREFAB_ROWS: list[tuple] = [
    (
        "UI",
        "MainMenuRoot.prefab",
        None,
        "Planned — not in project yet",
        "Main menu; entry point before gameplay",
        "Canvas for Play / Continue, Settings, Credits, Quit; optional background art slot.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "PauseMenuOverlay.prefab",
        None,
        "Planned — not in project yet",
        "In-game pause (time scale, input map)",
        "Resume, Settings shortcut, Quit to main menu; dimmed backdrop.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "GameHUD.prefab",
        None,
        "Planned — not in project yet",
        "Gameplay scenes (SampleScene, island, etc.)",
        "Crosshair or reticle, interaction prompts, objective text, stamina/oxygen if needed.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "SettingsPanel.prefab",
        None,
        "Planned — not in project yet",
        "Main menu and pause menu",
        "Tabs or pages: Audio (master/SFX/music), Graphics (quality, fullscreen), Controls (mouse sens), Language.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "LoadingScreen.prefab",
        None,
        "Planned — not in project yet",
        "Scene transitions / async loads",
        "Progress bar or spinner, optional gameplay tips, fade in/out.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "DialogueSubtitlesUI.prefab",
        None,
        "Planned — not in project yet",
        "Story beats, NPC interaction",
        "Speaker name, subtitle text, optional choice buttons; supports localization hooks.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "InventoryScreen.prefab",
        None,
        "Planned — not in project yet",
        "Full-screen or modal inventory (beyond hotbar)",
        "Grid or list of items, tooltips, sort/filter; wire to existing inventory system when ready.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "NotificationToast.prefab",
        None,
        "Planned — not in project yet",
        "Global UI layer",
        "Stacking toasts for quests, pickups, errors; auto-dismiss and queue.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
    (
        "UI",
        "CreditsOrAbout.prefab",
        None,
        "Planned — not in project yet",
        "Main menu → Credits",
        "Scrollable credits, version/build stamp, licenses link.",
        "—",
        "4 Wait",
        None,
        None,
        None,
        "Future placeholder",
    ),
]


def should_skip(dirpath: str) -> bool:
    return any(p in dirpath for p in SKIP_DIR_PARTS)


def load_guid_to_prefab() -> dict[str, str]:
    guid_to_prefab: dict[str, str] = {}
    for dirpath, _, filenames in os.walk(ROOT):
        if should_skip(dirpath):
            continue
        for fn in filenames:
            if not fn.endswith(".prefab.meta"):
                continue
            metapath = os.path.join(dirpath, fn)
            try:
                with open(metapath, "r", encoding="utf-8", errors="replace") as f:
                    m = GUID_RE.search(f.read())
            except OSError:
                continue
            if not m:
                continue
            prefab_path = metapath[:-5]
            rel = os.path.relpath(prefab_path, ROOT).replace(os.sep, "/")
            guid_to_prefab[m.group(1)] = rel
    return guid_to_prefab


def scan_prefab_guids(path: str) -> list[str]:
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
    except OSError:
        return []
    return SRC_PREFAB_RE.findall(text)


def load_editor_build_scenes() -> list[str]:
    """Return relative paths (posix) of enabled scenes in EditorBuildSettings."""
    p = os.path.join(ROOT, "ProjectSettings", "EditorBuildSettings.asset")
    if not os.path.isfile(p):
        return []
    try:
        with open(p, encoding="utf-8", errors="replace") as f:
            text = f.read()
    except OSError:
        return []
    paths: list[str] = []
    for m in re.finditer(
        r"-\s*enabled:\s*1\s*\r?\n\s*path:\s*(Assets/[^\s]+\.unity)",
        text,
    ):
        paths.append(m.group(1).replace(os.sep, "/"))
    return paths


def list_game_scene_paths() -> list[str]:
    """Project scenes under Assets/Scenes (game content roots)."""
    scenes_dir = os.path.join(ROOT, "Assets", "Scenes")
    out: list[str] = []
    if not os.path.isdir(scenes_dir):
        return out
    for dirpath, _, filenames in os.walk(scenes_dir):
        for fn in filenames:
            if fn.endswith(".unity"):
                full = os.path.join(dirpath, fn)
                out.append(os.path.relpath(full, ROOT).replace(os.sep, "/"))
    return sorted(out)


def collect_usage_from_game_scenes(
    guid_to_prefab: dict[str, str],
) -> dict[str, set[str]]:
    """
    Prefabs reachable from game scenes only: BFS from Assets/Scenes/*.unity and
    build-listed scenes, following nested prefab files.
    """
    roots: set[str] = set()
    roots.update(list_game_scene_paths())
    for p in load_editor_build_scenes():
        roots.add(p.replace(os.sep, "/"))

    queue: deque[str] = deque()
    seen_files: set[str] = set()
    for rel in sorted(roots):
        rel = rel.replace(os.sep, "/")
        full = os.path.join(ROOT, rel)
        if not rel.endswith(".unity") or not os.path.isfile(full):
            continue
        seen_files.add(full)
        queue.append(full)

    usage: dict[str, set[str]] = defaultdict(set)

    while queue:
        path = queue.popleft()
        rel_parent = os.path.relpath(path, ROOT).replace(os.sep, "/")
        for g in scan_prefab_guids(path):
            if g not in guid_to_prefab:
                continue
            usage[g].add(rel_parent)
            child_rel = guid_to_prefab[g].replace(os.sep, "/")
            if not child_rel.endswith(".prefab"):
                continue
            child_full = os.path.join(ROOT, child_rel)
            if not os.path.isfile(child_full):
                continue
            if child_full not in seen_files:
                seen_files.add(child_full)
                queue.append(child_full)

    return usage


def asset_type_for_path(rel: str) -> str:
    if rel.startswith("Assets/FruitAssets/"):
        return "Props"
    if rel.startswith("Assets/NatureAssets/"):
        return "Environment"
    if rel.startswith("Assets/GamedevDreamer/"):
        if "Controller" in rel or "Player" in rel:
            return "Character"
        return "Environment"
    if rel.startswith("Assets/Maritime_Heritage/"):
        return "Props"
    if rel.startswith("Assets/Boatassets/"):
        return "Vehicle"
    if rel.startswith("Assets/InventoryAsset/"):
        return "UI"
    if rel.startswith("Assets/Prefabs/"):
        return "Game"
    if rel.startswith("Assets/ElectricMotorAssets/"):
        return "Props"
    if rel.startswith("Assets/WaterPumpAssets/"):
        return "Props"
    if rel.startswith("Assets/SurvivalToolsAssets/"):
        return "Props"
    if rel.startswith("Assets/TextMesh Pro/"):
        return "UI"
    return "Prefab"


def format_context(refs: set[str]) -> str:
    scenes = sorted(r for r in refs if r.endswith(".unity"))
    nested = sorted(r for r in refs if r.endswith(".prefab"))
    parts: list[str] = []
    if scenes:
        names = [os.path.basename(s).replace(".unity", "") for s in scenes]
        parts.append("Scenes: " + ", ".join(names))
    if nested:
        short = [os.path.basename(p) for p in nested[:8]]
        extra = len(nested) - len(short)
        s = "Nested in: " + ", ".join(short)
        if extra > 0:
            s += f" (+{extra} more)"
        parts.append(s)
    return " | ".join(parts) if parts else ""


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "-o",
        "--output",
        default=os.path.join(
            os.path.expanduser("~"), "Downloads", "Asset List.xlsx"
        ),
        help="Output .xlsx path",
    )
    ap.add_argument(
        "--template",
        default="",
        help="Optional existing Asset List.xlsx to preserve header rows 1-9",
    )
    args = ap.parse_args()

    guid_to_prefab = load_guid_to_prefab()
    usage = collect_usage_from_game_scenes(guid_to_prefab)
    used = sorted(usage.keys(), key=lambda g: guid_to_prefab[g].lower())

    template_path = args.template
    if not template_path or not os.path.isfile(template_path):
        template_path = os.path.join(
            os.path.expanduser("~"), "Downloads", "Asset List.xlsx"
        )
    if os.path.isfile(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active

    # Keep first 9 rows if template has legend (header + status keys)
    header_row = 1
    data_start = 10
    if (
        ws.max_row >= 9
        and ws.cell(2, 1).value in (1, 1.0)
        and ws.cell(2, 9).value
    ):
        # Clear from row 10 down
        if ws.max_row >= data_start:
            ws.delete_rows(data_start, ws.max_row - data_start + 1)
    else:
        data_start = 1
        ws.append(
            [
                "ID",
                "Asset Type",
                "File Name",
                "Link to the Disk",
                "Path to the File in the Project",
                "Context/scene of use",
                "Description",
                "References",
                "Status",
                "Responsible",
                "Est, h",
                "To Date",
                "Comments",
            ]
        )
        header_row = 1
        data_start = 2

    for i, g in enumerate(used, start=1):
        rel = guid_to_prefab[g]
        name = os.path.basename(rel)
        refs = usage[g]
        row = data_start + i - 1
        ws.cell(row, 1, i)
        ws.cell(row, 2, asset_type_for_path(rel))
        ws.cell(row, 3, name)
        ws.cell(row, 4, None)
        ws.cell(row, 5, rel)
        ws.cell(row, 6, format_context(refs))
        ws.cell(
            row,
            7,
            "Prefab instance used under Assets/Scenes (and/or build settings scenes), "
            "including nested prefabs.",
        )
        ws.cell(row, 8, "; ".join(sorted(refs)))
        ws.cell(row, 9, "0 Done")
        ws.cell(row, 10, None)
        ws.cell(row, 11, None)
        ws.cell(row, 12, None)
        ws.cell(row, 13, None)

    base_id = len(used)
    next_row = data_start + len(used)
    for idx, cols in enumerate(FUTURE_PLANNED_PREFAB_ROWS, start=1):
        r = next_row + idx - 1
        ws.cell(r, 1, base_id + idx)
        for c, val in enumerate(cols, start=2):
            ws.cell(r, c, val)

    wb.save(args.output)
    print(
        f"Wrote {len(used)} prefab rows + {len(FUTURE_PLANNED_PREFAB_ROWS)} planned rows "
        f"to {args.output}"
    )


if __name__ == "__main__":
    main()
